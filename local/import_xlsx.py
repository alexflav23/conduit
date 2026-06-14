#!/usr/bin/env python3
"""Import the real H6Q workbook (the model finance actually maintains) straight into Conduit — no CSV step.

The workbook is a top-down forecast: a monthly P50 volume per *sales channel* (sheet "Monthly P50 Inc/Ex
Motability") plus a separate *SKU mix* (sheet "Overall Product Sales Mix": 5m/7.5m/10m x White/Black/Grey).
That is exactly the "agent inputs a unit count -> turn it into a per-SKU forecast via the historical mix" flow:
we read each channel's monthly volume, split it across the 9 SKUs by that channel's mix using the same
conserving largest-remainder allocation Conduit uses (sum of parts == the channel total, always), then sum to
a market-level per-SKU monthly forecast and emit the pipeline_coverage rows the board reads.

Usage:  import_xlsx.py <workbook.xlsx> [--ex-motability] [--year 2026]
Emits SQL on stdout (pipe into psql). "Inc Motability" -> the P50 base scenario; "Ex Motability" -> the
P50 ex_motability ex-cut scenario, matching the workbook's own toggle.
"""
import sys, argparse, datetime
from openpyxl import load_workbook

DEMO_MARKET = "22222222-2222-2222-2222-222222222222"
LENGTHS = ["5m", "7.5m", "10m"]
COLOURS = ["White", "Black", "Grey"]
# SKU code per (length, colour): HV-5M-W ... HV-10M-G. generation v3, one family "Hypervolt Charger".
SKU_CODE = {(l, c): f"HV-{l.upper().replace('.', '')}-{c[0]}" for l in LENGTHS for c in COLOURS}
# 9 mix cells laid out length-major (5m W/B/G, 7.5m W/B/G, 10m W/B/G) — the sheet's column order.
MIX_ORDER = [(l, c) for l in LENGTHS for c in COLOURS]
# Channels whose mix is the Octopus/Distributors weighted mix; everything else uses the retail mix.
WEIGHTED_CHANNELS = {"Hypervolt UK Energy", "Hypervolt UK Distributors", "Hypervolt UK Octopus"}


def largest_remainder(total, weights):
    """Allocate an integer total across weights so the parts sum to exactly total (conserving)."""
    s = sum(weights)
    if s <= 0 or total <= 0:
        return [0] * len(weights)
    raw = [total * w / s for w in weights]
    floors = [int(x) for x in raw]
    rem = total - sum(floors)
    order = sorted(range(len(weights)), key=lambda i: raw[i] - floors[i], reverse=True)
    for i in range(rem):
        floors[order[i]] += 1
    return floors


def find_header(rows):
    for idx, r in enumerate(rows):
        if len(r) > 1 and r[1] == "Sales Channel":
            return idx
    raise SystemExit("could not find the 'Sales Channel' header row")


def month_cols(header, year):
    out = {}
    for j, v in enumerate(header):
        if isinstance(v, datetime.datetime) and v.year == year:
            out[f"{v.year:04d}-{v.month:02d}"] = j
        elif isinstance(v, str) and v.startswith(f"{year}-") and len(v) == 7:
            out[v] = j
    if not out:
        raise SystemExit(f"no {year} monthly columns found in the header")
    return out


def read_mix(wb):
    rows = list(wb["Overall Product Sales Mix"].iter_rows(values_only=True))
    # The retail mix is the first all-numeric 9-wide row; the weighted mix is the next one.
    numeric = [r for r in rows if sum(1 for c in r[:9] if isinstance(c, (int, float))) >= 9]
    retail = [float(c) for c in numeric[0][:9]]
    weighted = [float(c) for c in numeric[1][:9]] if len(numeric) > 1 else retail
    return retail, weighted


def sql_str(s):
    return "'" + s.replace("'", "''") + "'"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("--ex-motability", action="store_true")
    ap.add_argument("--year", type=int, default=2026)
    ap.add_argument("--ndjson", action="store_true",
                    help="emit one JSON object per (sku, month) for the deterministic boot snapshot "
                         "(ingest/h6q/), instead of SQL — consumed by SnapshotLoader's h6q handler.")
    args = ap.parse_args()

    wb = load_workbook(args.workbook, read_only=True, data_only=True)
    sheet = "Monthly P50 Ex Motability" if args.ex_motability else "Monthly P50 Inc Motability"
    toggle = "'ex_motability'" if args.ex_motability else "NULL"
    if sheet not in wb.sheetnames:
        raise SystemExit(f"sheet '{sheet}' not in workbook")

    rows = list(wb[sheet].iter_rows(values_only=True))
    header = rows[find_header(rows)]
    months = month_cols(header, args.year)
    retail_mix, weighted_mix = read_mix(wb)

    # Accumulate market-level per-SKU monthly forecast = sum over top channels of (channel volume x mix).
    # The top channels are the 'x'-marked roll-up rows (Retail, Installers, Energy, Distributors, Automotive),
    # which sum to UK Total — using leaves would double-count.
    per_sku_month = {}  # (sku_code, month) -> qty
    channels_seen = []
    for r in rows:
        if not r or r[0] != "x" or not isinstance(r[1], str):
            continue
        name = r[1].strip()
        channels_seen.append(name)
        mix = weighted_mix if name in WEIGHTED_CHANNELS else retail_mix
        for month, col in months.items():
            v = r[col] if col < len(r) else None
            if not isinstance(v, (int, float)) or v <= 0:
                continue
            parts = largest_remainder(round(v), mix)
            for (ln, col_name), qty in zip(MIX_ORDER, parts):
                if qty:
                    key = (SKU_CODE[(ln, col_name)], month)
                    per_sku_month[key] = per_sku_month.get(key, 0) + qty

    skus = sorted({k[0] for k in per_sku_month})
    imported_months = sorted({k[1] for k in per_sku_month})

    if args.ndjson:
        import json
        toggle = "ex_motability" if args.ex_motability else None
        for (code, month) in sorted(per_sku_month):
            print(json.dumps({
                "sku": code,
                "period_month": month + "-01",
                "scenario": "P50",
                "toggle_basis": toggle,
                "forecast_qty": per_sku_month[(code, month)],
            }))
        sys.stderr.write(f"emitted {len(per_sku_month)} h6q coverage rows ({sum(per_sku_month.values())} units) as ndjson\n")
        return

    out = []
    out.append("BEGIN;")
    out.append("-- Real H6Q import from the finance workbook (channel volume x SKU mix -> per-SKU coverage).")
    out.append("INSERT INTO product_family (code, name) VALUES ('hv-charger','Hypervolt Charger') ON CONFLICT (code) DO NOTHING;")
    for ln in LENGTHS:
        for c in COLOURS:
            code = SKU_CODE[(ln, c)]
            name = f"Hypervolt Charger {ln} {c}"
            out.append(
                f"INSERT INTO product_variant (family_id, sku, generation) "
                f"SELECT id, {sql_str(code)}, 'v3' FROM product_family WHERE code='hv-charger' "
                f"ON CONFLICT (sku) DO NOTHING; -- {name}"
            )
    # Resolve the scenario once into a psql variable-free CTE-style subselect at insert time.
    out.append(
        "-- Wipe prior imported per-SKU market coverage for these months so re-running is idempotent."
    )
    month_list = ",".join(sql_str(m + "-01") for m in imported_months)
    sku_list = ",".join(sql_str(s) for s in skus)
    out.append(
        f"DELETE FROM pipeline_coverage pc USING product_variant pv "
        f"WHERE pc.level='market' AND pc.market_id={sql_str(DEMO_MARKET)} "
        f"AND pc.product_variant_id=pv.id AND pv.sku IN ({sku_list}) "
        f"AND pc.period_month IN ({month_list}) "
        f"AND pc.scenario_id=(SELECT id FROM forecast_scenario WHERE type='P50' AND toggle_basis IS NOT DISTINCT FROM {toggle} LIMIT 1);"
    )
    for (code, month) in sorted(per_sku_month):
        qty = per_sku_month[(code, month)]
        out.append(
            f"INSERT INTO pipeline_coverage (level, market_id, product_variant_id, period_month, scenario_id, forecast_qty) "
            f"SELECT 'market', {sql_str(DEMO_MARKET)}, pv.id, {sql_str(month + '-01')}, "
            f"(SELECT id FROM forecast_scenario WHERE type='P50' AND toggle_basis IS NOT DISTINCT FROM {toggle} LIMIT 1), {qty} "
            f"FROM product_variant pv WHERE pv.sku={sql_str(code)};"
        )
    out.append("COMMIT;")

    total = sum(per_sku_month.values())
    sys.stderr.write(
        f"Parsed {len(channels_seen)} channels {channels_seen}\n"
        f"-> {len(skus)} SKUs x {len(imported_months)} months ({args.year}"
        f"{', ex-motability' if args.ex_motability else ''}), {total} total units.\n"
    )
    print("\n".join(out))


if __name__ == "__main__":
    main()
